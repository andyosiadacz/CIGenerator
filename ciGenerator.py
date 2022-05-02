from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
import pandas as pd
import datetime as dt
import os, sys
from order import *


month = dt.date.today().month
day = dt.date.today().day
year = dt.date.today().year

file_date = f"{month}.{day}.{year}"
PROPER_DATE = f"{month}/{day}/{year}"


# ---- READ ORDERS FROM CSV INTO ORDER OBJECTS ----
def read_order_data(report):
    df = pd.read_csv(report, thousands=",")
    orders = []

    # ROW COUNTER
    row = 0
    

    order_num: int = None


    for num in df['Order Number']:
        row_data = df.loc[row]
        
        
        # CREATE NEW ORDER FOR NEW ORDER NUMBER AND LOAD HEADER DATA
        if order_num != num:

            order = Order(num)
            order_num = num

            # LOAD ORDER HEADER DATA
            order.sold_to_name = row_data['Sold To']
            order.sold_to_num = row_data['Sold To Number']
            order.po_num = row_data['Customer PO ']
            order.sold_to_address['Address Line 1'] = row_data["Address Line 1"]
            order.sold_to_address['Address Line 2'] = row_data["Address Line 2"]
            order.sold_to_address['City'] = row_data["City "]
            order.sold_to_address['State'] = row_data["ST "]
            order.sold_to_address['Postal Code'] = row_data["Postal Code"]
            order.sold_to_address['Country'] = row_data["Ctry "]

            orders.append(order)

        # INITIALIZE ORDERLINE, SAVE DATA FROM DF, AND APPEND ORDER LINE
        line = OrderLine()
        
        line.item_number = row_data['2nd Item Number']
        line.item_description = row_data['Concatenation Description']
        line.item_coo = "VN"
        line.item_hts_code = "8711.00.2600"
        line.item_price = row_data['Unit Price']
        line.quantity = row_data['Quantity Shipped']
        line.net_weight = 16.2
        line.gross_weight = 16.2
        line.cbm = 0.278

        order.orderlines.append(line)  

        # INCREMENT ROW   
        row += 1

    return orders


# ---- INVOICE CREATION FUNCTION DEFINITIONS ----
def create_ci_ca(order : Order):
    # --- LOAD TEMPLATE ---
    wb = load_workbook('templates/ca_ci_invoice_template.xlsx')
    ws = wb['Commercial Invoice']

    cinvoice_num =  f'{order.order_num} {file_date}'

    # ---- ADDRESS ----
    ws['C9'].value = order.sold_to_name
    ws['C10'].value = order.sold_to_address['Address Line 1']
    ws['C11'].value = f"{order.sold_to_address['City']}, {order.sold_to_address['State']} {order.sold_to_address['Postal Code']}"

    # ---- INVOICE DATA ----
    ws['G9'].value = cinvoice_num
    ws['G10'].value = file_date
    ws['G11'].value = order.po_num
    ws['G14'].value = order.sold_to_num
    ws['G17'].value = order.order_num

    # ---- ORDER LINE DATA ----
    row = 19
    for line in order.orderlines:
        ws[f'A{row}'].value = line.quantity
        ws[f'B{row}'].value = line.item_number
        ws[f'C{row}'].value = line.item_description
        ws[f'D{row}'].value = line.item_hts_code
        ws[f'E{row}'].value = line.item_coo
        ws[f'F{row}'].value = line.item_price
        ws[f"G{row}"].value = f'=A{row}*F{row}'
        
        row += 1

    #---- TAX RATE DATA ----
    if order.sold_to_address['State'] == "QC" or order.sold_to_address['State'] == "ON":
        ws['F51'].value = f"13% GST TAX"
        ws['G51'].value = f"=G48*0.13"

    #---- PACKING LIST DATA ---
    ws = wb['Packing List']

    row = 19
    for line in order.orderlines:
        ws[f'A{row}'].value = line.quantity
        ws[f'B{row}'].value = line.item_description
        ws[f'C{row}'].value = "EA"
        ws[f'D{row}'].value = line.net_weight
        ws[f'E{row}'].value = line.gross_weight
        ws[f'F{row}'].value = line.cbm

        row += 1

    #---- STATEMENT OF ORIGIN ----
    ws = wb['Statement of Origin']
    ws['A20'] = f"I certify that the goods described in this invoice or in the attached invoice #{cinvoice_num} were produced in the beneficiary country of {order.orderlines[0].item_coo}, and that at least 100% of the ex-factory price of goods originates in the beneficiary country of {order.orderlines[0].item_coo}."
    
    #----SAVE FILE ----
    try:
        os.mkdir(f'generated invoices')
        os.mkdir(f'generated invoices/{order.sold_to_name}')
        os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
    except FileExistsError:
        try:
            os.mkdir(f'generated invoices/{order.sold_to_name}')
            os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
        except FileExistsError:
            try:
                os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
            except:
                pass
    finally:
        wb.save(f'generated invoices/{order.sold_to_name}/{order.order_num}/CI {order.sold_to_name} {order.order_num} {file_date}.xlsx')
        print(f'Successfully created CI {order.sold_to_name} {order.order_num} {file_date}.xlsx')

def create_ci_exw(order: Order):
    #---- CELL STYLE VARIABLES ----
    lt_align = Alignment(horizontal='left')
    center_align = Alignment(horizontal='center')
    currency_format = "Currency"
    lt_border = Border(left=(Side(border_style='thick')))
    rt_border = Border(right=(Side(border_style='thick')))
    
    # --- LOAD TEMPLATE ---
    wb = load_workbook('templates/exw_ci_invoice_template.xlsx')
    ws = wb['CI']

    cinvoice_num =  f'{order.order_num} {file_date}'

    # ---- INVOICE DATA ----
    ws['N6'].value = cinvoice_num
    ws['F6'].value = PROPER_DATE


    # ---- ADDRESS ----
    ws['L10'].value = order.sold_to_name
    ws['L11'].value = order.sold_to_address['Address Line 1']
    ws['L12'].value = order.sold_to_address['Address Line 2']
    ws['L13'].value = f"{order.sold_to_address['City']}, {order.sold_to_address['Postal Code']}"
    ws['L14'].value = f"{order.sold_to_address['Country']}"


    # ---- LOAD & STYLE ORDER LINE DATA ----
    row = 24
    for line in order.orderlines:
        if row > 24:
            ws.insert_rows(row)

        ws[f'D{row}'].value = line.item_description
        ws[f'I{row}'].value = line.item_number
        ws[f'M{row}'].value = line.item_coo
        ws[f'O{row}'].value = line.item_hts_code
        ws[f'R{row}'].value = line.quantity
        ws[f'T{row}'].value = float(line.item_price)
        ws[f"V{row}"].value = f'=R{row}*T{row}'
        
        ws[f'D{row}'].alignment = lt_align
        ws[f'D{row}'].border = lt_border
        ws[f'I{row}'].alignment = center_align
        ws[f'O{row}'].alignment = center_align
        ws[f'R{row}'].alignment = center_align
        ws[f'V{row}'].border = rt_border

        row += 1


    #----SAVE FILE ----
    try:
        os.mkdir(f'generated invoices')
        os.mkdir(f'generated invoices/{order.sold_to_name}')
        os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
    except FileExistsError:
        try:
            os.mkdir(f'generated invoices/{order.sold_to_name}')
            os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
        except FileExistsError:
            try:
                os.mkdir(f'generated invoices/{order.sold_to_name}/{order.order_num}')
            except:
                pass
    finally:
        wb.save(f'generated invoices/{order.sold_to_name}/{order.order_num}/CI {order.sold_to_name} {order.order_num} {file_date}.xlsx')
        print(f'Successfully created CI {order.sold_to_name} {order.order_num} {file_date}.xlsx')

reports = sys.argv[1::]

for report in reports:
    orders = read_order_data(report)

for order in orders:
    if order.sold_to_address["Country"] == "CA":
        create_ci_ca(order)
    else:
        create_ci_exw(order)


